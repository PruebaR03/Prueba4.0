import os
import pandas as pd
from openpyxl import load_workbook
from ..core import limpiar_ruta, leer_instrucciones, leer_excel_o_csv
from ..core.file_utils import asegurar_carpeta, aplicar_formato_encabezados, colorear_pestanas_resumen, reordenar_hojas_por_categoria
from .filtros import aplicar_criterio, aplicar_criterios_multiples

def procesar_archivo(ruta: str, instruccion: dict) -> pd.DataFrame | None:
    df = leer_excel_o_csv(ruta)
    if df is None:
        return None

    df.columns = [str(c).strip().strip('"').strip("'").strip().lower() for c in df.columns]
    
    columna_id = instruccion.get("columna id", "").strip().strip('"').strip("'").strip().lower()
    limpiar_id_patron = instruccion.get("limpiar id", "").strip().strip('"').strip("'")

    print(f"  📊 Columnas disponibles: {', '.join(df.columns.tolist()[:10])}...")

    if columna_id and columna_id not in df.columns:
        print(f"  ❌ Columna ID '{columna_id}' no existe")
        print(f"  💡 Columnas disponibles completas: {list(df.columns)}")
        return None

    tiene_multiples_criterios = 'columna criterio 2' in instruccion or 'criterio 2' in instruccion
    
    columna_criterio_principal = instruccion.get("columna criterio", "").strip().strip('"').strip("'").strip()
    criterio_principal = instruccion.get("criterio", "").strip()
    
    sin_filtros = not columna_criterio_principal or not criterio_principal
    
    if tiene_multiples_criterios:
        i = 2
        while f'columna criterio {i}' in instruccion or f'criterio {i}' in instruccion:
            col_crit = instruccion.get(f'columna criterio {i}', '').strip().strip('"').strip("'").strip()
            crit = instruccion.get(f'criterio {i}', '').strip()
            if col_crit and crit:
                sin_filtros = False
                break
            i += 1
    
    if sin_filtros:
        print("  ℹ️  Sin criterios de filtrado válidos → Incluyendo todas las filas")
        print(f"  ✅ Total: {len(df)} fila(s)\n")
        df_filtrado = df
    elif tiene_multiples_criterios:
        df_filtrado, _ = aplicar_criterios_multiples(df, instruccion)
    else:
        columna_criterio = columna_criterio_principal.lower()
        
        if columna_criterio and columna_criterio not in df.columns:
            print(f"  ⚠️  Columna criterio '{columna_criterio}' no existe → Incluyendo todas las filas")
            df_filtrado = df
        else:
            df_filtrado, _ = aplicar_criterio(df, columna_criterio, criterio_principal)
    
    if columna_id and limpiar_id_patron:
        print(f"  🧹 Limpiando extensión '{limpiar_id_patron}' de columna '{columna_id}'")
        df_filtrado[columna_id] = df_filtrado[columna_id].apply(
            lambda x: str(x).replace(limpiar_id_patron, "").strip() if pd.notna(x) else x
        )
        print(f"     ✅ Limpieza completada")
    
    return df_filtrado

def generar_excel_salida(ruta_instrucciones: str, ruta_salida: str):
    ruta_instrucciones = limpiar_ruta(ruta_instrucciones)
    ruta_salida = limpiar_ruta(ruta_salida)
    asegurar_carpeta(ruta_salida)

    instrucciones = leer_instrucciones(ruta_instrucciones)
    hojas_creadas = False

    try:
        print("\n" + "═" * 70)
        print("🚀 INICIANDO GENERACIÓN DE EXCEL BASE")
        print("═" * 70)
        
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            for idx, instruccion in enumerate(instrucciones, 1):
                nombre_hoja = instruccion.get("archivo") or f"Hoja_{idx}"
                ruta = instruccion.get("ruta", "").strip()

                print(f"\n📄 [{idx}/{len(instrucciones)}] Procesando: {nombre_hoja}")
                print("─" * 70)

                if ruta:
                    df_filtrado = procesar_archivo(ruta, instruccion)
                    if df_filtrado is not None:
                        df_filtrado.to_excel(writer, sheet_name=nombre_hoja, index=False)
                        hojas_creadas = True
                        
                        if df_filtrado.empty:
                            print(f"  ⚠️  Hoja '{nombre_hoja}' creada vacía (sin coincidencias)")
                        else:
                            print(f"  ✅ Hoja '{nombre_hoja}' creada con {len(df_filtrado)} fila(s)")
                else:
                    pd.DataFrame().to_excel(writer, sheet_name=nombre_hoja, index=False)
                    print(f"  ℹ️  Hoja '{nombre_hoja}' creada vacía (sin ruta configurada)")
                    hojas_creadas = True

            if not hojas_creadas:
                df_vacio = pd.DataFrame({"Mensaje": ["No se generaron datos válidos"]})
                df_vacio.to_excel(writer, sheet_name="Hoja_Predeterminada", index=False)

        print(f"\n🎨 Aplicando formato visual...")
        wb = load_workbook(ruta_salida)
        aplicar_formato_encabezados(wb)
        wb.save(ruta_salida)
        wb.close()
        
        print("\n" + "═" * 70)
        print(f"✅ ARCHIVO EXCEL GENERADO EXITOSAMENTE")
        print(f"📁 Ubicación: {ruta_salida}")
        print("═" * 70 + "\n")
    except PermissionError:
        print(f"\n❌ ERROR: No se pudo escribir en '{ruta_salida}'")
        print("💡 Solución: Cierre el archivo Excel si está abierto.\n")
    except Exception as e:
        print(f"\n❌ ERROR INESPERADO: {e}\n")

def crear_hoja_resumen(ruta_excel: str, ruta_instrucciones: str):
    configuraciones = leer_instrucciones(ruta_instrucciones)
    ruta_excel = limpiar_ruta(ruta_excel)

    if not os.path.exists(ruta_excel):
        print(f"❌ Excel '{ruta_excel}' no existe. No se puede crear resumen.")
        return

    try:
        print("\n" + "═" * 70)
        print("📊 CREANDO HOJA RESUMEN")
        print("═" * 70 + "\n")
        
        with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            excel = pd.ExcelFile(ruta_excel)
            
            categorias = {}
            sin_categoria = []
            
            for configuracion in configuraciones:
                categoria = configuracion.get("categoria", "").strip()
                if categoria:
                    if categoria not in categorias:
                        categorias[categoria] = []
                    categorias[categoria].append(configuracion)
                else:
                    sin_categoria.append(configuracion)
            
            if not categorias:
                print("  ℹ️  No se encontraron categorías → Creando resumen consolidado único")
                _crear_resumen_simple(excel, configuraciones, writer, "Resumen")
            else:
                print(f"  📁 {len(categorias)} categoría(s) detectada(s):")
                for cat in categorias.keys():
                    print(f"     • {cat}")
                
                for categoria, configs in categorias.items():
                    print(f"\n  📋 Procesando categoría: {categoria}")
                    nombre_hoja = f"Resumen_{categoria.replace(' ', '_')}"
                    _crear_resumen_categoria(excel, configs, writer, nombre_hoja, categoria)
                
                if sin_categoria:
                    print(f"\n  📋 Procesando hojas sin categoría")
                    _crear_resumen_categoria(excel, sin_categoria, writer, "Resumen_General", "General")
                
                print(f"\n  ℹ️  Con categorías definidas → No se crea resumen consolidado")

        wb = load_workbook(ruta_excel)
        
        for ws_name in wb.sheetnames:
            if ws_name.startswith("Resumen") or ws_name == "Resumen":
                print(f"  🎨 Formateando resumen: {ws_name}")
                aplicar_formato_encabezados(wb, ws_name)
        
        colorear_pestanas_resumen(wb)
        
        if categorias:
            categorias_por_hoja_map = {}
            for categoria, configs in categorias.items():
                for config in configs:
                    hoja = config.get("archivo")
                    if hoja:
                        categorias_por_hoja_map[hoja] = categoria
            
            reordenar_hojas_por_categoria(wb, categorias_por_hoja_map)
            print(f"  🔄 Hojas reordenadas: resúmenes al final de cada categoría")
        
        wb.save(ruta_excel)
        wb.close()

        print("\n" + "═" * 70)
        print("✅ HOJA(S) RESUMEN CREADA(S) EXITOSAMENTE")
        print("═" * 70 + "\n")

    except Exception as e:
        print(f"\n❌ Error al crear hoja resumen: {e}\n")
        import traceback
        traceback.print_exc()

def _crear_resumen_simple(excel: pd.ExcelFile, configuraciones: list, writer, nombre_hoja: str = "Resumen"):
    resumen = {}
    
    for idx, configuracion in enumerate(configuraciones, 1):
        hoja = configuracion.get("archivo")
        columna_id = configuracion.get("columna id")

        if not hoja or not columna_id or hoja not in excel.sheet_names:
            continue

        print(f"     🔍 [{idx}/{len(configuraciones)}] Analizando: {hoja}")
        
        df_hoja = excel.parse(hoja)
        df_hoja.columns = [str(c).strip().strip('"').strip("'").strip().lower() for c in df_hoja.columns]
        columna_id = columna_id.strip().strip('"').strip("'").strip().lower()

        if columna_id not in df_hoja.columns or df_hoja.empty:
            print(f"        ⚠️  Sin datos válidos")
            continue

        ids = df_hoja[columna_id].dropna().unique()
        print(f"        ✅ {len(ids)} ID(s) únicos encontrados")
        
        for id_val in ids:
            id_str = str(id_val).strip()
            if id_str not in resumen:
                resumen[id_str] = {}
            resumen[id_str][hoja] = "X"

    if resumen:
        df_resumen = pd.DataFrame.from_dict(resumen, orient="index").fillna("")
        df_resumen.index.name = "ID"
        df_resumen.reset_index(inplace=True)
    else:
        df_resumen = pd.DataFrame(columns=["ID"])

    if len(df_resumen.columns) > 1:
        otras_cols = sorted([c for c in df_resumen.columns if c != "ID"])
        df_resumen = df_resumen[["ID"] + otras_cols]

    df_resumen.to_excel(writer, sheet_name=nombre_hoja, index=False)
    print(f"     📋 {nombre_hoja}: {len(df_resumen)} registro(s)")

def _crear_resumen_categoria(excel: pd.ExcelFile, configuraciones: list, writer, nombre_hoja: str, categoria: str):
    resumen = {}
    hojas_procesadas = []
    
    for configuracion in configuraciones:
        hoja = configuracion.get("archivo")
        columna_id = configuracion.get("columna id")

        if not hoja or not columna_id or hoja not in excel.sheet_names:
            continue

        hojas_procesadas.append(hoja)
        
        df_hoja = excel.parse(hoja)
        df_hoja.columns = [str(c).strip().strip('"').strip("'").strip().lower() for c in df_hoja.columns]
        columna_id = columna_id.strip().strip('"').strip("'").strip().lower()

        if columna_id not in df_hoja.columns or df_hoja.empty:
            continue

        ids = df_hoja[columna_id].dropna().unique()
        
        for id_val in ids:
            id_str = str(id_val).strip()
            if id_str not in resumen:
                resumen[id_str] = {}
            resumen[id_str][hoja] = "X"

    if resumen:
        df_resumen = pd.DataFrame.from_dict(resumen, orient="index").fillna("")
        df_resumen.index.name = "ID"
        df_resumen.reset_index(inplace=True)
    else:
        df_resumen = pd.DataFrame(columns=["ID"])

    if len(df_resumen.columns) > 1:
        otras_cols = sorted([c for c in df_resumen.columns if c != "ID"])
        df_resumen = df_resumen[["ID"] + otras_cols]

    df_resumen.to_excel(writer, sheet_name=nombre_hoja, index=False)
    print(f"     ✅ {len(hojas_procesadas)} hoja(s) procesadas → {len(df_resumen)} registro(s)")
