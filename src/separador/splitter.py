import os
import pandas as pd
from openpyxl import load_workbook
from ..core import limpiar_ruta, leer_configuracion_separacion, leer_instrucciones
from ..core.file_utils import aplicar_formato_encabezados, colorear_pestanas_resumen, reordenar_hojas_por_categoria
from ..enriquecedor.enricher import _enriquecer_desde_archivo, _preparar_cache_lookup

def celda_contiene_identificador(valor, identificadores: list) -> bool:
    """
    Verifica si celda coincide con identificador (maneja N/A, vacíos, etc).
    """
    identificadores_lower = [str(i).lower().strip() for i in identificadores]
    
    if "n/a" in identificadores_lower:
        if pd.isna(valor) or valor is None:
            return True
        valor_str = str(valor).strip().upper()
        if valor_str in ["N/A", "NA", ""]:
            return True
    
    if pd.isna(valor) or valor is None:
        return False
    
    v = str(valor).lower().strip()
    if not v:
        return "" in identificadores_lower
    
    for ident in identificadores_lower:
        if ident in v:
            return True
    
    return False

def _map_id_por_hoja_desde_instrucciones(ruta_instrucciones: str) -> dict:
    """
    Mapea hoja -> columna id desde instrucciones.
    """
    try:
        instrucciones = leer_instrucciones(ruta_instrucciones)
    except Exception as e:
        print(f"⚠ No se pudo leer instrucciones: {e}")
        return {}
    
    mapa = {}
    for inst in instrucciones:
        hoja = inst.get("archivo")
        col_id = inst.get("columna id")
        if hoja and col_id:
            mapa[str(hoja)] = str(col_id).strip().lower()
    return mapa

def _map_categorias_desde_instrucciones(ruta_instrucciones: str) -> dict:
    """
    Mapea hoja -> categoría desde instrucciones.
    """
    try:
        instrucciones = leer_instrucciones(ruta_instrucciones)
    except Exception as e:
        print(f"⚠ No se pudo leer instrucciones: {e}")
        return {}
    
    mapa = {}
    for inst in instrucciones:
        hoja = inst.get("archivo")
        categoria = inst.get("categoria", "").strip()
        if hoja:
            mapa[str(hoja)] = categoria if categoria else "General"
    return mapa

def flujo_separacion(ruta_excel: str, ruta_configuracion: str, carpeta_salida: str = "output/separados", ruta_instrucciones: str | None = None):
    """
    Separa Excel en múltiples archivos según identificadores.
    Crea resúmenes por categoría en cada archivo separado.
    COPIA HOJAS YA ENRIQUECIDAS del archivo base.
    """
    ruta_excel = limpiar_ruta(ruta_excel)
    config = leer_configuracion_separacion(ruta_configuracion)

    if not config['hojas_calculo'] or not config['hojas']:
        print("❌ Configuración inválida o incompleta.")
        return

    if not os.path.exists(ruta_excel):
        print(f"❌ Excel base no existe: {ruta_excel}")
        return

    id_por_hoja = _map_id_por_hoja_desde_instrucciones(ruta_instrucciones) if ruta_instrucciones else {}
    categorias_por_hoja = _map_categorias_desde_instrucciones(ruta_instrucciones) if ruta_instrucciones else {}
    carpeta_salida = limpiar_ruta(carpeta_salida)
    os.makedirs(carpeta_salida, exist_ok=True)

    try:
        print("\n" + "═" * 70)
        print("✂️  INICIANDO SEPARACIÓN DE ARCHIVOS EXCEL")
        print("═" * 70)
        print(f"📂 Carpeta destino: {carpeta_salida}")
        print(f"📋 Archivos a generar: {len(config['hojas_calculo'])}")

        excel_base = pd.ExcelFile(ruta_excel)
        hojas_existentes = set(excel_base.sheet_names)
        
        # Verificar hojas que están en el Excel pero no en la configuración
        hojas_configuradas = set(h['hoja'] for h in config['hojas'])
        hojas_no_configuradas = hojas_existentes - hojas_configuradas - {h for h in hojas_existentes if h.startswith("Resumen")}
        
        if hojas_no_configuradas:
            print(f"\n⚠️  ADVERTENCIA: Hojas en el Excel base que NO están en la configuración:")
            for hoja in sorted(hojas_no_configuradas):
                print(f"     ❌ '{hoja}'")
            print(f"\n💡 Estas hojas NO serán incluidas en los archivos separados.")
            print(f"💡 Agrégalas a '{os.path.basename(ruta_configuracion)}' si deseas incluirlas.\n")

        for idx, calc in enumerate(config['hojas_calculo'], 1):
            nombre_salida = calc['name']
            identificadores = calc.get('identificadores', [])
            
            if not identificadores:
                print(f"\n  ⚠️  '{nombre_salida}' sin identificadores → Omitido")
                continue

            print(f"\n" + "─" * 70)
            print(f"📄 [{idx}/{len(config['hojas_calculo'])}] Generando: {nombre_salida}")
            print(f"   🏷️  Identificadores: {', '.join(identificadores)}")
            print("─" * 70)

            ruta_out = os.path.join(carpeta_salida, f"{nombre_salida}.xlsx")

            with pd.ExcelWriter(ruta_out, engine='openpyxl') as writer:
                # Resúmenes por categoría
                resumenes_por_categoria = {}
                total_filas = 0

                for hoja_cfg in config['hojas']:
                    hoja = hoja_cfg.get('hoja')
                    col_sep = hoja_cfg.get('columna_id', '').strip()
                    
                    if not hoja or not col_sep or hoja not in hojas_existentes:
                        continue

                    # ✅ LEER HOJA YA ENRIQUECIDA (con todas sus columnas)
                    df = excel_base.parse(hoja)
                    
                    # Guardar columnas originales antes de normalizar
                    columnas_originales = list(df.columns)
                    
                    # Normalizar solo para búsqueda (crear columnas temporales)
                    df_cols_norm = {str(c).strip().strip('"').strip("'").strip().lower(): c for c in df.columns}
                    col_sep_norm = col_sep.strip().strip('"').strip("'").strip().lower()
                    
                    if col_sep_norm not in df_cols_norm:
                        print(f"     ❌ Hoja '{hoja}': Columna '{col_sep}' no existe")
                        continue

                    # Usar columna original (no normalizada) para filtrar
                    columna_original = df_cols_norm[col_sep_norm]
                    
                    # Filtrar por identificadores
                    mask = df[columna_original].apply(lambda v: celda_contiene_identificador(v, identificadores))
                    df_filtrado = df[mask]

                    if df_filtrado.empty:
                        # Guardar con nombres de columnas originales
                        pd.DataFrame(columns=columnas_originales).to_excel(writer, sheet_name=hoja, index=False)
                        print(f"     ⚠️  Hoja '{hoja}': 0 filas (vacía)")
                    else:
                        # Guardar con columnas originales (ya enriquecidas)
                        df_filtrado.to_excel(writer, sheet_name=hoja, index=False)
                        total_filas += len(df_filtrado)
                        print(f"     ✅ Hoja '{hoja}': {len(df_filtrado)} fila(s) (ENRIQUECIDAS)")

                        # Determinar categoría de la hoja
                        categoria = categorias_por_hoja.get(hoja, "General")
                        if categoria not in resumenes_por_categoria:
                            resumenes_por_categoria[categoria] = {}

                        # Columna para resumen (usar normalizada para mapeo)
                        col_resumen_norm = id_por_hoja.get(hoja, col_sep_norm)
                        col_resumen_original = df_cols_norm.get(col_resumen_norm, columna_original)
                        
                        # Acumular IDs por categoría
                        if "n/a" in [str(i).lower() for i in identificadores]:
                            ids_validos = df_filtrado[col_resumen_original].unique()
                        else:
                            ids_validos = df_filtrado[col_resumen_original].dropna().unique()
                        
                        for id_val in ids_validos:
                            key = "N/A" if pd.isna(id_val) else str(id_val).strip()
                            if key not in resumenes_por_categoria[categoria]:
                                resumenes_por_categoria[categoria][key] = {}
                            resumenes_por_categoria[categoria][key][hoja] = "X"

                # Crear hojas de resumen por categoría
                categorias_unicas = set(categorias_por_hoja.values()) if categorias_por_hoja else set()
                
                if len(categorias_unicas) > 1:
                    # Hay múltiples categorías, crear resúmenes separados
                    print(f"\n     📊 Creando resúmenes por categoría:")
                    for categoria, resumen_cat in resumenes_por_categoria.items():
                        nombre_resumen = f"Resumen_{categoria.replace(' ', '_')}"
                        
                        if resumen_cat:
                            df_resumen = pd.DataFrame.from_dict(resumen_cat, orient="index").fillna("")
                            df_resumen.index.name = "ID"
                            df_resumen.reset_index(inplace=True)
                            

                            # Ordenar columnas: ID primero, luego alfabéticamente
                            if len(df_resumen.columns) > 1:
                                otras_cols = sorted([c for c in df_resumen.columns if c.lower() != "id"])
                                df_resumen = df_resumen[["ID"] + otras_cols]
                        else:
                            df_resumen = pd.DataFrame(columns=["ID"])
                        
                        df_resumen.to_excel(writer, sheet_name=nombre_resumen, index=False)
                        print(f"        ✅ {nombre_resumen}: {len(df_resumen)} ID(s)")
                else:
                    # Una sola categoría o sin categorías, crear resumen único
                    print(f"\n     📊 Creando resumen único:")
                    resumen_consolidado = {}
                    for resumen_cat in resumenes_por_categoria.values():
                        for id_key, hojas_dict in resumen_cat.items():
                            if id_key not in resumen_consolidado:
                                resumen_consolidado[id_key] = {}
                            resumen_consolidado[id_key].update(hojas_dict)
                    
                    if resumen_consolidado:
                        df_resumen = pd.DataFrame.from_dict(resumen_consolidado, orient="index").fillna("")
                        df_resumen.index.name = "ID"
                        df_resumen.reset_index(inplace=True)
                        
                        # Ordenar columnas: ID primero, luego alfabéticamente
                        if len(df_resumen.columns) > 1:
                            otras_cols = sorted([c for c in df_resumen.columns if c.lower() != "id"])
                            df_resumen = df_resumen[["ID"] + otras_cols]
                    else:
                        df_resumen = pd.DataFrame(columns=["ID"])
                    
                    df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
                    print(f"        ✅ Resumen: {len(df_resumen)} ID(s)")

            # Aplicar formato
            wb = load_workbook(ruta_out)
            aplicar_formato_encabezados(wb)
            
            # Colorear pestañas de resumen
            colorear_pestanas_resumen(wb)
            
            # Reordenar hojas si hay múltiples categorías
            if len(categorias_unicas) > 1 and categorias_por_hoja:
                reordenar_hojas_por_categoria(wb, categorias_por_hoja)
                print(f"        🔄 Hojas reordenadas con resúmenes al final de cada categoría")
            
            wb.save(ruta_out)
            wb.close()
            
            # Enriquecer resúmenes del archivo separado si existe ruta de Excel base
            if ruta_excel:
                _enriquecer_resumenes_archivo_separado(ruta_out, ruta_excel)
            
            print(f"\n   💾 Archivo guardado: {nombre_salida}.xlsx")
            print(f"   📈 Total de filas: {total_filas}")

        print("\n" + "═" * 70)
        print("✅ SEPARACIÓN COMPLETADA EXITOSAMENTE")
        print(f"📁 Archivos generados en: {carpeta_salida}")
        print("═" * 70 + "\n")

    except Exception as e:
        print(f"\n❌ Error durante la separación: {e}")
        import traceback
        traceback.print_exc()

def _enriquecer_resumenes_archivo_separado(ruta_archivo_separado: str, ruta_excel_base: str):
    """
    Enriquece las hojas resumen de un archivo separado usando el Excel base como fuente.
    MATCHEA POR CATEGORÍA: Solo aplica el enriquecimiento de la categoría correspondiente.
    """
    try:
        # Leer Excel base para obtener datos de enriquecimiento
        excel_base = pd.ExcelFile(ruta_excel_base)
        
        # Mapear resúmenes base por categoría
        resumenes_base_por_categoria = {}
        
        for sheet_name in excel_base.sheet_names:
            if sheet_name.startswith("Resumen_"):
                # Extraer categoría del nombre (Resumen_Workstations → Workstations)
                categoria = sheet_name.replace("Resumen_", "").replace("_", " ")
                
                df_temp = excel_base.parse(sheet_name)
                df_temp.columns = [str(c).strip().strip('"').strip("'").strip().lower() for c in df_temp.columns]
                
                # Buscar columnas de enriquecimiento (todas excepto ID y nombres de hojas)
                columnas_adicionales = []
                for col in df_temp.columns:
                    if col == "id":
                        continue
                    # Columnas de enriquecimiento: contienen espacios o son metadatos conocidos
                    if " " in col or col in ["employee id", "assigned to", "department", "location", "status", "status reason"]:
                        columnas_adicionales.append(col)
                
                if columnas_adicionales:
                    resumenes_base_por_categoria[categoria] = {
                        "df": df_temp,
                        "columnas": columnas_adicionales
                    }
        
        if not resumenes_base_por_categoria:
            # No hay resúmenes enriquecidos en el Excel base
            return
        
        print(f"\n     🔗 Enriqueciendo resúmenes del archivo separado por categoría...")

        # Abrir archivo separado y enriquecer sus resúmenes
        with pd.ExcelWriter(ruta_archivo_separado, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            excel_separado = pd.ExcelFile(ruta_archivo_separado)
            
            hojas_enriquecidas = 0
            for sheet_name in excel_separado.sheet_names:
                if not sheet_name.startswith("Resumen"):
                    continue
                
                # Extraer categoría del resumen separado
                if sheet_name.startswith("Resumen_"):
                    categoria_separado = sheet_name.replace("Resumen_", "").replace("_", " ")
                else:
                    # Resumen sin categoría específica
                    categoria_separado = None
                
                # Buscar resumen base correspondiente
                resumen_base_info = None
                if categoria_separado and categoria_separado in resumenes_base_por_categoria:
                    resumen_base_info = resumenes_base_por_categoria[categoria_separado]
                elif not categoria_separado:
                    # Si es un resumen genérico, intentar usar cualquiera (fallback)
                    if resumenes_base_por_categoria:
                        resumen_base_info = list(resumenes_base_por_categoria.values())[0]
                
                if not resumen_base_info:
                    print(f"        ⚠️  '{sheet_name}': Sin resumen base correspondiente para categoría '{categoria_separado}'")
                    continue
                
                df_resumen_base = resumen_base_info["df"]
                columnas_validas = resumen_base_info["columnas"]
                
                # Leer resumen del archivo separado
                df_resumen = excel_separado.parse(sheet_name)
                df_resumen.columns = [str(c).strip().strip('"').strip("'").strip().lower() for c in df_resumen.columns]
                
                if "id" not in df_resumen.columns:
                    continue
                
                # Crear diccionario de mapeo ID -> datos de enriquecimiento
                mapeo_enriquecimiento = {}
                for _, row in df_resumen_base.iterrows():
                    id_val = row["id"]
                    if pd.notna(id_val):
                        id_str = str(id_val).strip()
                        mapeo_enriquecimiento[id_str] = {}
                        for col in columnas_validas:
                            if col in df_resumen_base.columns and pd.notna(row[col]):
                                mapeo_enriquecimiento[id_str][col] = row[col]
                
                if not mapeo_enriquecimiento:
                    continue
                
                # Agregar columnas de enriquecimiento si no existen
                posicion_insercion = 1  # Después de ID
                for col in columnas_validas:
                    if col not in df_resumen.columns:
                        df_resumen.insert(posicion_insercion, col, "N/A")
                        posicion_insercion += 1
                
                # Enriquecer con datos
                ids_enriquecidos = 0
                for idx, row in df_resumen.iterrows():
                    id_val = str(row["id"]).strip()
                    if id_val in mapeo_enriquecimiento:
                        for col, valor in mapeo_enriquecimiento[id_val].items():
                            df_resumen.at[idx, col] = valor
                        ids_enriquecidos += 1
                
                # Guardar hoja enriquecida
                df_resumen.to_excel(writer, sheet_name=sheet_name, index=False)
                hojas_enriquecidas += 1
                print(f"        ✅ '{sheet_name}': {ids_enriquecidos} ID(s) enriquecidos con {len(columnas_validas)} columna(s)")
                print(f"           📊 Columnas: {', '.join(columnas_validas)}")
        
        if hojas_enriquecidas > 0:
            # Reaplicar formato después del enriquecimiento
            wb = load_workbook(ruta_archivo_separado)
            for ws_name in wb.sheetnames:
                if ws_name.startswith("Resumen"):
                    aplicar_formato_encabezados(wb, ws_name)
            wb.save(ruta_archivo_separado)
            print(f"        🎨 Formato aplicado a resúmenes enriquecidos")
            
    except Exception as e:
        print(f"        ⚠️  No se pudo enriquecer resúmenes: {e}")
        import traceback
        traceback.print_exc()
