import os
from openpyxl.styles import Font, PatternFill, Border, Side

def limpiar_ruta(ruta: str) -> str:
    """
    Limpia rutas eliminando prefijos file:/// y caracteres especiales.
    """
    if ruta is None:
        return ""
    
    # Eliminar comillas dobles y simples del inicio y final
    ruta = ruta.strip().strip('"').strip("'")
    
    # Eliminar comillas internas que quedaron (como "valor")
    ruta = ruta.replace('"', '').replace("'", '')
    
    # Procesar prefijos file://
    if ruta.startswith("file:///"):
        ruta = ruta[8:].replace("%20", " ")
    elif ruta.startswith("file:"):
        ruta = ruta[5:].replace("%20", " ")
    
    return ruta

def asegurar_carpeta(ruta: str):
    """
    Crea la carpeta si no existe.
    """
    carpeta = os.path.dirname(ruta)
    if carpeta and not os.path.exists(carpeta):
        os.makedirs(carpeta)

def aplicar_formato_encabezados(workbook, sheet_name: str = None):
    """
    Aplica formato rojo de fondo y texto blanco a los encabezados, y bordes negros a TODAS las celdas CON DATOS.
    
    Args:
        workbook: Objeto Workbook de openpyxl
        sheet_name: Nombre de hoja específica o None para todas
    """
    # Estilos para encabezados: fondo rojo, texto blanco en negrita
    font_blanco = Font(color="FFFFFF", bold=True)
    fill_rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Estilo de bordes negros
    borde_delgado = Side(style='thin', color='000000')
    borde_completo = Border(
        left=borde_delgado,
        right=borde_delgado,
        top=borde_delgado,
        bottom=borde_delgado
    )
    
    hojas = [workbook[sheet_name]] if sheet_name else workbook.worksheets
    
    for ws in hojas:
        # Calcular RANGO REAL de datos (sin filas vacías al final)
        max_row_real = ws.max_row
        max_col_real = ws.max_column
        
        # Buscar la última fila con datos reales
        for row_idx in range(ws.max_row, 0, -1):
            tiene_datos = False
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip() != "":
                    tiene_datos = True
                    break
            if tiene_datos:
                max_row_real = row_idx
                break
        
        # Buscar la última columna con datos reales
        for col_idx in range(ws.max_column, 0, -1):
            tiene_datos = False
            for row_idx in range(1, max_row_real + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip() != "":
                    tiene_datos = True
                    break
            if tiene_datos:
                max_col_real = col_idx
                break
        
        if max_row_real == 0 or max_col_real == 0:
            continue
        
        print(f"     🎨 Aplicando formato a '{ws.title}': {max_row_real} filas x {max_col_real} columnas (rango real)")
        
        # 1. Aplicar formato a TODOS los encabezados (primera fila)
        for col_idx in range(1, max_col_real + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_blanco
            cell.fill = fill_rojo
            cell.border = borde_completo
        
        # 2. Aplicar bordes SOLO a celdas con datos reales
        for row_idx in range(1, max_row_real + 1):
            for col_idx in range(1, max_col_real + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = borde_completo

def colorear_pestanas_resumen(workbook):
    """
    Aplica color rojo tenue (FF6B6B) a las pestañas de hojas que empiezan con 'Resumen'.
    
    Args:
        workbook: Objeto Workbook de openpyxl
    """
    for ws in workbook.worksheets:
        if ws.title.startswith("Resumen") or ws.title == "Resumen":
            # Color rojo tenue
            ws.sheet_properties.tabColor = "FF6B6B"

def reordenar_hojas_por_categoria(workbook, categorias_por_hoja: dict):
    """
    Reordena las hojas para que las hojas resumen vayan al final de cada categoría.
    
    Args:
        workbook: Objeto Workbook de openpyxl
        categorias_por_hoja: Dict mapeando nombre_hoja -> categoria
    """
    # Agrupar hojas por categoría
    hojas_por_categoria = {}
    hojas_resumen = []
    hojas_sin_categoria = []
    
    for ws in workbook.worksheets:
        nombre = ws.title
        
        if nombre.startswith("Resumen"):
            hojas_resumen.append(ws)
        elif nombre in categorias_por_hoja:
            categoria = categorias_por_hoja[nombre]
            if categoria not in hojas_por_categoria:
                hojas_por_categoria[categoria] = []
            hojas_por_categoria[categoria].append(ws)
        else:
            hojas_sin_categoria.append(ws)
    
    # Reordenar: hojas normales por categoría + resumen de esa categoría
    nuevo_orden = []
    categorias_procesadas = set()
    
    for categoria, hojas in hojas_por_categoria.items():
        # Agregar hojas normales de la categoría
        nuevo_orden.extend(hojas)
        categorias_procesadas.add(categoria)
        
        # Buscar y agregar resumen de esta categoría
        nombre_resumen = f"Resumen_{categoria.replace(' ', '_')}"
        for resumen in hojas_resumen:
            if resumen.title == nombre_resumen:
                nuevo_orden.append(resumen)
                break
    
    # Agregar hojas sin categoría
    nuevo_orden.extend(hojas_sin_categoria)
    
    # Agregar resúmenes restantes al final
    for resumen in hojas_resumen:
        if resumen not in nuevo_orden:
            nuevo_orden.append(resumen)
    
    # Aplicar nuevo orden
    for idx, ws in enumerate(nuevo_orden):
        workbook.move_sheet(ws, offset=idx - workbook.index(ws))
