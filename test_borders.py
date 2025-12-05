from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def test_bordes(ruta_excel: str):
    """Prueba de aplicación de bordes."""
    wb = load_workbook(ruta_excel)
    ws = wb.active
    
    borde = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    print(f"Aplicando bordes a {ws.max_row} filas x {ws.max_column} columnas")
    
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = borde
            if row_idx == 1 and col_idx == 1:
                print(f"Celda A1 border: {cell.border}")
    
    wb.save(ruta_excel)
    print("✅ Bordes aplicados")

# Ejecutar
test_bordes("output/tu_archivo.xlsx")
